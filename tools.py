#!/user/bin/env python
# coding=utf-8
"""
@project : ImageManager
@ide     : PyCharm
@file    : main
@author  : wuhoubo
@desc    :
@create  : 2021/11/13 15:57:59
@update  :
"""
import os
import random
import re
import time
from collections import Counter

from bs4 import BeautifulSoup
import requests
import json
from openpyxl import load_workbook
from PIL import Image, ImageFile

from helper.db_helper import DBHelper, DBExecuteType
from helper.file_helper import FileHelper
from helper.image_helper import ImageHelper
from model.data import MyImage
from PIL import Image, ImageFile

from tools.baidu_translate import baidu_translate

ImageFile.LOAD_TRUNCATED_IMAGES = True
Image.MAX_IMAGE_PIXELS = None


def analysis_and_rename_file(dir_path, prefix, handler):
    for root, dirs, files in os.walk(dir_path):
        length = len(files)
        for i in range(length):
            file = files[i]
            file_path = os.path.join(root, file)
            print(f'[{i}/{length}] {file_path}')
            handler(file_path, prefix)


db_helper = DBHelper(None)


def update_image(file_path, prefix):
    info = ImageHelper.analyze_image_info(file_path)
    source = info.source
    if source == 'pixiv':
        sub_str = f'_{info.tags}'
    elif source in ['yande', 'konachan']:
        if '00000' in info.relative_path:
            sub_str = f'{info.tags}_00000'
        elif '_00' in info.relative_path:
            sub_str = f'{info.tags}_00'
        else:
            sub_str = f'{info.tags}'
    else:
        sub_str = ''
    new_path = file_path.replace(sub_str, '')
    if file_path == new_path:
        print('未匹配或已经优化过文件名，直接跳过')
        return
    try:
        if not os.path.exists(new_path):
            os.rename(file_path, new_path)
    except Exception as e:
        print(f"重命名失败：{e}")
        return
    md5 = FileHelper.get_md5(new_path)
    old = db_helper.search_by_md5(md5)
    if not old:
        print('没有历史数据，跳过')
        return
    new_path = new_path.replace('\\', '/')
    image = MyImage(id=old.id, desc=info.desc, author=info.author, type=old.type, level=old.type,
                    tags=info.tags, works=old.works, role=old.role, source=info.source, width=old.width,
                    height=old.height, size=old.size, relative_path=new_path.replace(prefix, ''), md5=md5,
                    file_create_time=FileHelper.get_create_time_str(new_path), series=old.series,
                    uploader=info.uploader, sequence=old.sequence)
    db_helper.update_image(image)


def recheck_size(start_page):
    page_size = 500
    error_f = open('error.log', 'a+', encoding='utf-8')
    not_exist_f = open('notExist.log', 'a+', encoding='utf-8')
    count = db_helper.get_table_count("select count(*) from myacg.image;")
    while True:
        offset = (start_page - 1) * page_size
        sql = f"select * from image limit {page_size} offset {offset};"
        queries = db_helper.execute(sql, DBExecuteType.FetchAll)
        if not queries:
            print('没有数据，结束检查')
            break
        infos = [MyImage.from_mysql_dict(x) for x in queries]
        n = len(infos)
        for i in range(n):
            info = infos[i]
            print(f'[{offset + i}/{count}] {info.id}, {info.relative_path}')
            if not os.path.exists(info.path):
                not_exist_f.write(f'{info.id},{info.path}\n')
                print('图片不存在')
                db_helper.delete(info.id)
                continue
            try:
                width, height = Image.open(info.path).size
            except Exception as e:
                error_f.write(f'{info.id},{info.relative_path},{info.width},{info.height},{e}\n')
                print('尺寸读取失败')
                continue
            if not width or not height:
                error_f.write(f'{info.id},{info.relative_path},{info.width},{info.height}\n')
                print('尺寸读取失败')
                continue
            if width == info.width and height == info.height:
                continue
            print(f'宽度：{info.width} -> {width}，高度：{info.height} -> {height}')
            info.width = width
            info.height = height
            db_helper.update_image(info)
        start_page += 1
    error_f.close()
    not_exist_f.close()


def check_no_record_image(file_path, prefix):
    if not ImageHelper.is_image(file_path):
        return
    relative_path = file_path.replace('\\', '/').replace(prefix, '')
    info = db_helper.search_by_file_path(relative_path)
    if not info:
        with open('notRecord.log', 'a+', encoding='utf-8') as f:
            f.write(f'{file_path}\n')
        print('文件不存在')


def rename_png2jpg(dir_path):
    for root, dirs, files in os.walk(dir_path):
        length = len(files)
        for i in range(length):
            file = files[i]
            file_path = os.path.join(root, file)
            print(f'[{i}/{length}] {file_path}')
            extension = FileHelper.get_file_extension(file_path)
            if extension != '.png':
                continue
            new_path = file_path.replace('.png', '.jpg')
            os.rename(file_path, new_path)


def get_file_paths(dir_path):
    paths = []
    img_extension = ['jpg', 'jpeg', 'bmp', 'png']
    # img_extension = ['png']
    for root, dirs, files in os.walk(dir_path):
        for file in files:
            if file.lower().split('.')[-1] in img_extension:
                paths.append([root, os.path.join(root, file)])
    return paths


def compress(source_dir, target_dir):
    paths = get_file_paths(source_dir)
    n = len(paths)
    total_percent = 0
    for i in range(n):
        path = os.path.join(paths[i][1])
        img = Image.open(path)
        if img.mode != 'RGB':
            img = img.convert('RGB')
        filename = path.split('/')[-1].split('.')[0]
        target_sub_dir = os.path.join(target_dir, paths[i][0].replace(source_dir, ''))
        new_path = os.path.join(target_sub_dir, filename + '.jpg')
        # if os.path.exists(new_path):
        #     continue
        no = 1
        while os.path.exists(new_path):
            new_path = os.path.join(target_sub_dir, f'{filename}_{no:0>2d}.jpg')
            print(f'有重复，重命名为：{new_path}')
            no += 1
        if not os.path.exists(target_sub_dir):
            print(f'创建文件夹{target_sub_dir}')
            os.makedirs(target_sub_dir)
        img.save(new_path, quality=90)
        old_size = FileHelper.get_file_size_in_mb(path)
        new_size = FileHelper.get_file_size_in_mb(new_path)
        sub = old_size - new_size
        percent = round(new_size * 100 / old_size, 2)
        total_percent += percent
        print(f'[{i}/{n}]压缩比例 {percent:.2f}%, 减小大小 {round(sub / old_size, 2):.2f} MB, {path.replace(source_dir, "")}')
    print(f'总压缩比：{round(total_percent / n, 2):.2f}%')


def check_exist(file_path, prefix):
    if not ImageHelper.is_image(file_path):
        return
    md5 = FileHelper.get_md5(file_path)
    info = db_helper.search_by_md5(md5)
    if info:
        print('已存在，删除该文件')
        os.remove(file_path)


def get_html(url, cookies=None):
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/64.0.3282.119 Safari/537.36'}
    proxies = {"http": "http://127.0.0.1:7890", "https": 'http://127.0.0.1:7890'}
    return requests.get(url, headers=headers, proxies=proxies, timeout=12, cookies=cookies).text


def get_not_exist_yande_tag():
    imgs, _ = db_helper.search_by_where("source='yande' and uploader=''")
    n = len(imgs)
    i = 0
    while i < n:
        img = imgs[i]
        no = ImageHelper.get_yande_no(img.relative_path)
        if not no:
            print(f'[{i}/{n}]{img.id} - {img.relative_path}')
            i += 1
            continue
        url = f'https://yande.re/post/show/{no}'
        try:
            html = get_html(url)
        except Exception as e:
            print(f'失败：{e}')
            time.sleep(2)
            continue
        val = BeautifulSoup(html, 'lxml')
        ul = val.find('div', id='stats').find('ul')
        li = ul.li.next_sibling.next_sibling
        uploader = li.a.next_sibling.next_sibling
        if not uploader:
            uploader = 'Anonymous'
        else:
            uploader = uploader.get_text()
        print(f'[{i}/{n}]{img.id} - {uploader} - {img.relative_path}')
        db_helper.execute(f"update image set uploader='{uploader}' where id={img.id}", DBExecuteType.Run)
        duration = random.uniform(0, 2)
        print(f'休眠 {duration}s')
        time.sleep(duration)
        i += 1


def get_not_exist_pixiv_tag():
    imgs, _ = db_helper.search_by_where("source='pixiv' and tags=''")
    n = len(imgs)
    i = 0
    while i < n:
        img = imgs[i]
        no = ImageHelper.get_pixiv_no(img.relative_path)
        if not no:
            print(f'[{i}/{n}]找不到 no {img.id} - {img.relative_path}')
            i += 1
            continue
        cookies = {}
        with open('cookies.txt') as f:
            obj = json.loads(f.read())
            for item in obj:
                cookies[item['name']] = item['value']
        try:
            html = get_html(f'https://www.pixiv.net/artworks/{no}', cookies)
        except Exception as e:
            print(f'失败：{e}')
            time.sleep(3)
            continue
        if '该作品已被删除，或作品ID不存在。' in html:
            print(f'[{i}/{n}]作品不存在 {img.id} - {img.relative_path}')
            i += 1
            continue
        val = BeautifulSoup(html, 'lxml')
        meta = val.find('meta', id='meta-preload-data')
        obj = json.loads(meta.attrs['content'])
        print(json.dumps(obj))
        break
        obj_tags = obj['illust'][no]['tags']['tags']
        tags = []
        for tag in obj_tags:
            tag_str = tag['tag']
            if 'translation' in tag:
                tag_str += f"({tag['translation']['en']})"
            tags.append(tag_str)
        tag_str = ';'.join(tags)
        tag_str = tag_str.replace("'", "\\'")
        print(f'[{i}/{n}]{img.id} - {no},{img.desc} - {tag_str}')
        img.tags = tag_str
        db_helper.execute(f"update image set tags='{tag_str}' where id={img.id}", DBExecuteType.Run)
        duration = random.uniform(0, 2)
        print(f'休眠 {duration}s')
        time.sleep(duration)
        i += 1


def refresh_yande_tag():
    workbook = load_workbook(filename='yande.xlsx')
    sheet = workbook.worksheets[1]
    tag_di, role_di, works_di, author_di, company_di = {}, {}, {}, {}, {}
    type2di = {'作品': works_di, '角色': role_di, '作者': author_di, '公司': company_di, '标签': tag_di}
    for row in sheet.iter_rows(min_row=2):
        trans = row[2].value
        if not trans:
            continue
        source = row[1].value
        types = row[3].value
        if not types:
            tag_di[source] = trans
            continue
        if types == '系列':
            types = '作品'
        if types in ['原画', '画师', '鬼刀']:
            types = '作者'
        if types in type2di:
            type2di[types][source] = trans
            continue
        if types in ['作品&标签', '不确定']:
            continue
        trans = trans.split(';')
        if '；' in types:
            types = types.split('；')
        else:
            types = types.split(';')
        if len(trans) != len(types):
            continue
        type2di[types[0]][source] = trans[0]
        type2di[types[1]][source] = trans[1]
    workbook.close()

    with open('record/tag.log', 'w+') as f:
        for k, v in tag_di.items():
            f.write(f'{k},{v}\n')
    with open('record/role.log', 'w+') as f:
        for k, v in role_di.items():
            f.write(f'{k},{v}\n')
    with open('record/works.log', 'w+') as f:
        for k, v in works_di.items():
            f.write(f'{k},{v}\n')
    with open('record/author.log', 'w+') as f:
        for k, v in author_di.items():
            f.write(f'{k},{v}\n')
    with open('record/company.log', 'w+') as f:
        for k, v in company_di.items():
            f.write(f'{k},{v}\n')

    return

    imgs, _ = db_helper.search_by_where("source='yande' and tags!=''")
    n = len(imgs)
    i = 0
    while i < n:
        img = imgs[i]
        tags = img.tags.split(' ')
        new_tags = []
        for tag in tags:
            if tag in tag_di:
                new_tags.append(tag_di[tag])
            elif tag in role_di:
                img.role = role_di[tag]
            elif tag in works_di:
                img.works = works_di[tag]
            elif tag in author_di:
                img.author = author_di[tag]
            elif tag in company_di:
                img.author = company_di[tag]
            else:
                new_tags.append(tag)

        print(f'[{i}/{n}]{img.id} - {img.relative_path}')
        tag_str = ';'.join(new_tags).replace("'", "\\'")
        author = img.author.replace("'", "\\'")
        works = img.works.replace("'", "\\'")
        db_helper.execute(
            f"update image set role='{img.role}',works='{works}',author='{author}',tags='{tag_str}' where id={img.id}",
            DBExecuteType.Run)
        i += 1


def get_not_tran_yande_tag():
    imgs, _ = db_helper.search_by_where("source='yande' and tags!=''")
    tags = []
    pattern = re.compile(r'[\u4e00-\u9fa5]')
    allow_tags = ['photoshop', 'CG', '2D', 'Cosplay', 'tagme', 'hololive']
    for img in imgs:
        tp = img.tags.split(';')
        tags += [x for x in tp if not re.search(pattern, x) and x not in allow_tags]
    counter = Counter(tags)
    counter = counter.most_common()
    # sorted(counter)
    lines = ['数量,原名,翻译名,类型\n']
    n = len(counter)
    for i in range(n):
        k, v = counter[i]
        tran = baidu_translate(k)
        line = f"{v},{k},{tran},\n"
        print(f'[{i}/{n}] {v},{k},{tran}')
        lines.append(line)
    with open('yande.csv', 'w+') as f:
        f.writelines(lines)


if __name__ == '__main__':
    # analysis_and_rename_file(r'Z:\写真', 'Z:/', check_no_record_image)
    # rename_png2jpg(r'E:\下载\第四资源站\楚楚子 png待功能完善')
    refresh_yande_tag()
