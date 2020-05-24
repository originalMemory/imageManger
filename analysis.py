#!/user/bin/env python
# coding=utf-8
"""
@project : ImageManager
@ide     : PyCharm
@file    : main
@author  : wuhoubo
@desc    : 
@create  : 2020/5/11 8:39:45
@update  :
"""

from googletrans import Translator
from openpyxl import Workbook
from openpyxl import load_workbook
from helper.db_helper import DBHelper
from json import JSONDecodeError
from random import randint
import time


def extractYandeTags():
    wb = Workbook()
    sheet = wb.get_active_sheet()
    sheet.append(('数量', '标签'))
    source = 'yande'
    offset = 0
    db_helper = DBHelper(error_handler)
    tags = {}
    while True:
        sql_str = f"select `desc` from image where source='{source}' limit 500 offset {offset};"
        print(sql_str)
        queries = db_helper.query_with_return_all(sql_str)
        if not queries:
            break

        for query in queries:
            tag_str = query['desc']
            if not tag_str:
                continue
            local_tags = str(tag_str).split(' ')
            for tag in local_tags:
                if tag in tags:
                    count = tags[tag]
                else:
                    count = 0
                tags[tag] = count + 1
        offset += 500

    tags = sorted(tags.items(), key=lambda x: x[1], reverse=True)
    for (tag, count) in tags:
        sheet.append((count, tag))
    wb.save("yande标签列表.xlsx")


def error_handler(error_str):
    print(f"数据库错误：{error_str}")


def translate(startRow):
    filename = "yande标签列表.xlsx"
    wb = load_workbook(filename)
    sheet = wb.get_active_sheet()
    tags = []
    translator = Translator(service_urls=['translate.google.cn'])
    count = 300
    while startRow < sheet.max_row:
        print(f"翻译 {startRow}/{sheet.max_row}")
        for i in range(0, count):
            currentRow = startRow + i
            if currentRow > sheet.max_row:
                break
            tags.append(sheet.cell(currentRow, 2).value)
        src = "，".join(tags)
        try:
            translates = translator.translate(src, dest='zh-CN',src='en')
            dests = translates.text.split("，")
            for index, text in enumerate(dests):
                print(f"[{startRow + index}/{sheet.max_row}] ", tags[index], ' -> ', text)
                sheet.cell(startRow + index, 3).value = text
            time.sleep(random())
        except JSONDecodeError as error:
            print("到达上限")
            break
        startRow += count
        tags.clear()

    wb.save(filename)

def random():
  random_number =randint(0, 10)
  return random_number

if __name__ == '__main__':
    translate(2)
